import base64
from os import getcwd
from io import BytesIO
from odoo import models, fields
import openpyxl


class ImportWizard(models.TransientModel):

    _name = 'employee.wizard.import'

    import_employees = fields.Binary(string="Import Employees")

    def create_emp_fun(self):
        print(self.id)
        binary = self.env["ir.attachment"].browse(self._context.get("active_ids")).search(
            [("res_model", "=", "employee.wizard.import"), ("res_id", "=", self.id), ("res_field", "=", "import_employees")])
        path = binary.store_fname
        print(path)
        file_datas = binary._full_path(binary.store_fname)
        print('file_datas', file_datas)
        dataframe = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.import_employees)))
        # Define variable to read sheet
        dataframe1 = dataframe.active
        # Iterate the loop to read the cell values
        for row in range(1, dataframe1.max_row):
            employees = []
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)
                employees.append(col[row].value)
            print(employees)
            self.env['hr.employee'].create({'name': employees[0], 'work_email': employees[1], 'work_phone': employees[2]})
